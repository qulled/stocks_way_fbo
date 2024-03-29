from logging.handlers import RotatingFileHandler

from dotenv import load_dotenv
from googleapiclient import discovery
from google.oauth2 import service_account
from googleapiclient.discovery import build
import logging
import os
import datetime as dt

import openpyxl
import warnings

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(BASE_DIR, 'logs/')
log_file = os.path.join(BASE_DIR, 'logs/pars_stocks_table.log')
console_handler = logging.StreamHandler()
file_handler = RotatingFileHandler(
    log_file,
    maxBytes=100000,
    backupCount=3,
    encoding='utf-8'
)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s, %(levelname)s, %(message)s',
    handlers=(
        file_handler,
        console_handler
    )
)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
CREDENTIALS_FILE = 'credentials_service.json'
credentials = service_account.Credentials.from_service_account_file(CREDENTIALS_FILE)
service = discovery.build('sheets', 'v4', credentials=credentials)
START_POSITION_FOR_PLACE = 1

dotenv_path = os.path.join(os.path.dirname(__file__), '.env')

if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)
load_dotenv('.env ')
SPREADSHEET_ID = os.getenv('SPREADSHEET_ID')


def list_barcode(employees_sheet):
    list_barcode = []
    for x in range(2, employees_sheet.max_row + 1):
        article = str(employees_sheet.cell(row=x, column=8).value)
        if article not in list_barcode and article.isnumeric():
            list_barcode.append(str(employees_sheet.cell(row=x, column=8).value))
    return list_barcode


def dict_price(table_id):
    dict_price = {}
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    for item in values[1:]:
        dict_price[str(item[4].upper())] = 0
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range='Цены закупок', majorDimension='ROWS').execute()
    values = result.get('values', [])
    for item in values[1:]:
        for article in dict_price:
            if article.startswith(item[0]):
                dict_price[article]=item[1]
                break
    return dict_price


def dicts_info(employees_sheet, list_barcode):
    dict_ur_lico = {}
    dict_brand = {}
    dict_subject = {}
    dict_nomenclature = {}
    dict_article = {}
    dict_size = {}
    for x in range(2, employees_sheet.max_row + 1):
        article = str(employees_sheet.cell(row=x, column=8).value)
        if article.isnumeric():
            article = str(employees_sheet.cell(row=x, column=8).value)
        if article in list_barcode:
            dict_ur_lico[article] = employees_sheet.cell(row=x, column=3).value
        if article in list_barcode:
            dict_brand[article] = employees_sheet.cell(row=x, column=4).value
        if article in list_barcode:
            dict_subject[article] = employees_sheet.cell(row=x, column=5).value
        if article in list_barcode:
            dict_article[article] = employees_sheet.cell(row=x, column=6).value
        if article in list_barcode:
            dict_nomenclature[article] = employees_sheet.cell(row=x, column=7).value
        if article in list_barcode:
            dict_size[article] = employees_sheet.cell(row=x, column=9).value
    return dict_ur_lico, dict_brand, dict_subject, dict_article, dict_nomenclature, dict_size


def dicts_way(employees_sheet):
    dict_way_to_from_client = {}
    for x in range(2, employees_sheet.max_row + 1):
        article = str(employees_sheet.cell(row=x, column=8).value)
        if article not in dict_way_to_from_client:
            dict_way_to_from_client[article] = employees_sheet.cell(row=x,
                                                                    column=10).value
    return dict_way_to_from_client


def convert_to_column_letter(column_number):
    column_letter = ''
    while column_number != 0:
        c = ((column_number - 1) % 26)
        column_letter = chr(c + 65) + column_letter
        column_number = (column_number - c) // 26
    return column_letter


def update_table_barcode(table_id, list_barcode):
    position_for_place = START_POSITION_FOR_PLACE

    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                index = 0
                if len(list_barcode) != 0:
                    value = list_barcode[index]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 5)}{i}',
                         'values': [[f'{value}']]}]
                    list_barcode.pop(index)
            except Exception as e:
                print(e)
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_ur_lico(table_id, dict_ur_lico):
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                barcode = row[5].strip().upper()
                if barcode in dict_ur_lico:
                    value = dict_ur_lico[barcode]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place)}{i}',
                         'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_brand(table_id, dict_brand):
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                barcode = row[5].strip().upper()
                if barcode in dict_brand:
                    value = dict_brand[barcode]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 1)}{i}',
                         'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_category(table_id, dict_subject):
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                barcode = row[5].strip().upper()
                if barcode in dict_subject:
                    value = dict_subject[barcode]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 2)}{i}',
                         'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_nomenclature(table_id, dict_nomenclature):
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                barcode = row[5].strip().upper()
                if barcode in dict_nomenclature:
                    value = dict_nomenclature[barcode]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 3)}{i}',
                         'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_article(table_id, dict_article):
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                barcode = row[5].strip().upper()
                if barcode in dict_article:
                    value = dict_article[barcode]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 4)}{i}',
                         'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_size(table_id, dict_size):
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                barcode = row[5].strip().upper()
                if barcode in dict_size:
                    value = dict_size[barcode]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 6)}{i}',
                         'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_prise(dict_price):
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                article = row[4].strip().upper()
                if article in dict_price:
                    if str(dict_price[article]).isnumeric():
                        if int(dict_price[article]) > 0:
                            value = dict_price[article]
                            body_data += [
                                {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 7)}{i}',
                                 'values': [[f'{value}']]}]
                    else:
                        value = ''
                        body_data += [
                            {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 7)}{i}',
                             'values': [[f'{value}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


def update_table_to_client(table_id, dict_way):
    position_for_place = START_POSITION_FOR_PLACE
    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()
    values = result.get('values', [])
    i = 2
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[1:]:
            try:
                barcode = row[5].strip().upper()
                ex = str(dict_way[barcode])
                if barcode in dict_way and ex.isnumeric():
                    value = dict_way[barcode]
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 8)}{i}',
                         'values': [[f'{value}']]}]
                else:
                    value = ''
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 8)}{i}',
                         'values': [[f'{value}']]}]
            except Exception as e:
                continue
                # print(e)
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


if __name__ == '__main__':
    date_from = dt.datetime.date(dt.datetime.now())
    range_name = 'В пути'
    table_id = SPREADSHEET_ID
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        excel_file = openpyxl.load_workbook(f'final_excel/ALL-{dt.datetime.date(dt.datetime.now())}.xlsx')
    employees_sheet = excel_file['Sheet1']
    dict_ur_lico, dict_brand, dict_subject, dict_article, dict_nomenclature, dict_size = \
        dicts_info(employees_sheet, list_barcode(employees_sheet))
    update_table_barcode(table_id, list_barcode(employees_sheet))
    update_table_ur_lico(table_id, dict_ur_lico)
    update_table_brand(table_id, dict_brand)
    update_table_category(table_id, dict_subject)
    update_table_nomenclature(table_id, dict_nomenclature)
    update_table_article(table_id, dict_article)
    update_table_size(table_id, dict_size)
    update_table_prise(dict_price(table_id))
    update_table_to_client(table_id, dicts_way(employees_sheet))
